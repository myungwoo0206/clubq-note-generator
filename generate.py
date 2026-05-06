from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from html import escape
import os
import re

EXCEL_FILE = "data.xlsx"


# =====================
# 기본 유틸
# =====================

def clean(value):
    return "" if value is None else str(value).strip()


def normalize(value):
    return clean(value).replace("\n", "").replace(" ", "")


def apply_highlight(value):
    """
    엑셀에서 *강조*로 입력한 부분을 노란색 + 볼드로 변환
    예: *루닛 이수스페셜티케미컬*
    """
    if value is None:
        return ""

    text = escape(str(value))

    text = re.sub(
        r"\*(.+?)\*",
        r'<span class="important-red">\1</span>',
        text
    )

    return text.replace("\n", "<br>")


def split_lines_text(text):
    lines = []

    for line in clean(text).splitlines():
        line = line.strip()
        line = re.sub(r"^[-•]\s*", "", line)
        if line:
            lines.append(line)

    return lines


def parse_market_value(text):
    text = clean(text)
    match = re.match(r"([0-9,.]+)\s*\(([^)]+)\)", text)

    if match:
        return match.group(1), match.group(2)

    return text, ""


def is_up(change):
    change = clean(change)

    if change.startswith("-"):
        return False

    number = re.sub(r"[^0-9.\-]", "", change)

    try:
        return float(number) >= 0
    except:
        return True


def format_theme_desc(raw_text):
    """
    06번 테마 영역:
    - 추세강화), 강세), 신규) 라벨은 주황색 유지
    - 뒤의 설명은 흰색
    - *강조*는 노란색 + 볼드
    - 줄바꿈 유지
    """
    lines = clean(raw_text).splitlines()
    formatted = []

    for line in lines:
        raw = line.strip()
        if not raw:
            continue

        match = re.match(r"^(추세강화\)|강세\)|신규\))\s*(.*)", raw)

        if match:
            label = escape(match.group(1))
            body = apply_highlight(match.group(2))

            formatted.append(
                f'<div class="theme-line">'
                f'<span class="theme-label">{label}</span> '
                f'<span class="theme-text">{body}</span>'
                f'</div>'
            )
        else:
            formatted.append(
                f'<div class="theme-line">'
                f'<span class="theme-text">{apply_highlight(raw)}</span>'
                f'</div>'
            )

    return "".join(formatted)


# =====================
# 엑셀 파싱
# =====================

def parse_excel():
    workbook = load_workbook(EXCEL_FILE, data_only=True)
    sheet = workbook.active

    note = {
        "date": "",
        "thoughts": [],
        "stocks": [],
        "kospi": {"value": "", "change": ""},
        "kosdaq": {"value": "", "change": ""},
        "clubq": [],
        "trades": [],
        "themes": [],
    }

    header = clean(sheet["A1"].value)
    if header:
        note["date"] = header.splitlines()[0].strip()

    current_section = ""

    for row in sheet.iter_rows(min_row=3):
        a, b, c = (list(row) + [None, None, None])[:3]

        section_raw = clean(a.value)
        if section_raw:
            current_section = normalize(section_raw)

        name_text = clean(b.value)
        content_text = clean(c.value)

        name_html = apply_highlight(b.value)
        content_html = apply_highlight(c.value)

        if not current_section:
            continue

        if current_section == "시장에대한생각":
            source = name_text or content_text
            note["thoughts"].extend(split_lines_text(source))

        elif current_section == "상한가등특징주":
            if name_text:
                note["stocks"].append({
                    "name": name_html,
                    "desc": content_html or "-"
                })

        elif current_section == "시장":
            if name_text == "코스피":
                value, change = parse_market_value(content_text)
                note["kospi"] = {"value": value, "change": change}

            elif name_text == "코스닥":
                value, change = parse_market_value(content_text)
                note["kosdaq"] = {"value": value, "change": change}

        elif current_section == "ClubQ모임소식":
            source = name_text or content_text
            for line in split_lines_text(source):
                note["clubq"].append(escape(line))

        elif current_section in [
            "호스트의실제매매및관심기업(업종)",
            "호스트의실제매매및관심기업"
        ]:
            if name_text:
                note["trades"].append({
                    "type": name_html,
                    "desc": content_html
                })

        elif current_section in [
            "주도업종을찾기위한업종흐름파악",
            "업종흐름"
        ]:
            if name_text:
                note["themes"].append({
                    "week": name_html,
                    "desc": format_theme_desc(content_text)
                })

    return note


note = parse_excel()

kospi_img = "up_green.png" if is_up(note["kospi"]["change"]) else "down_green.png"
kosdaq_img = "up_blue.png" if is_up(note["kosdaq"]["change"]) else "down_blue.png"


# =====================
# HTML 행 생성
# =====================

def rows_stock(stocks):
    return "".join([
        f"""
        <div class="table-row">
            <div class="stock-name">{item['name']}</div>
            <div class="stock-desc">{item['desc']}</div>
        </div>
        """
        for item in stocks
    ])


def rows_trades(trades):
    return "".join([
        f"""
        <div class="trade-row">
            <div class="trade-type">{item['type']}</div>
            <div class="trade-desc">{item['desc']}</div>
        </div>
        """
        for item in trades
    ])


def rows_themes(themes):
    return "".join([
        f"""
        <div class="theme-row">
            <div class="theme-week">{item['week']}</div>
            <div class="theme-desc">{item['desc']}</div>
        </div>
        """
        for item in themes
    ])


thoughts_html = "".join([
    f"<li>{escape(item)}</li>"
    for item in note["thoughts"]
])

clubq_html = "<br>".join(note["clubq"])


# =====================
# HTML 템플릿
# =====================

html_content = f"""
<html>
<head>
<meta charset="utf-8">
<style>
* {{
    box-sizing: border-box;
}}

html, body {{
    margin: 0;
    padding: 0;
    background: #020b12;
    font-family: 'NanumGothic', 'Nanum Gothic', 'Apple SD Gothic Neo', Arial, sans-serif;
    color: white;
    width: fit-content;
    height: fit-content;
}}

.note {{
    width: 1350px;
    display: inline-block;
    padding: 34px;
    background:
        radial-gradient(circle at top right, rgba(0,206,209,0.18), transparent 28%),
        linear-gradient(180deg, #03131c 0%, #020b12 100%);
}}

.important-red {{
    color: #FFD400;
    font-weight: 900;
}}

.header {{
    display: grid;
    grid-template-columns: 380px 320px 1fr;
    gap: 34px;
    align-items: center;
    border-bottom: 2px solid #00CED1;
    padding-bottom: 24px;
}}

.date {{
    color: #00F0E0;
    font-size: 44px;
    font-weight: 900;
}}

.title {{
    font-size: 50px;
    font-weight: 900;
    white-space: nowrap;
    line-height: 1.05;
}}

.link {{
    font-size: 30px;
    color: #8ffdfa;
    border-left: 2px solid white;
    padding-left: 30px;
    white-space: nowrap;
    font-weight: 800;
}}

.slogan {{
    font-size: 30px;
    font-weight: 800;
    line-height: 1.4;
    border-left: 2px solid white;
    padding-left: 30px;
}}

.slogan span {{
    color: #00CED1;
}}

.section {{
    display: grid;
    grid-template-columns: 300px 1fr;
    border: 1.5px solid #00AEB5;
    border-radius: 16px;
    overflow: hidden;
    margin-top: 14px;
    background: rgba(0, 30, 42, 0.78);
}}

.section-left {{
    padding: 22px;
    border-right: 1px solid rgba(0,206,209,0.45);

    display: grid;
    grid-template-columns: 72px 1fr;
    align-items: center;
    column-gap: 18px;
}}

.num {{
    width: 58px;
    height: 48px;
    background: #00AEB5;
    color: white;
    font-size: 30px;
    font-weight: 900;
    border-radius: 9px;

    display: flex;
    align-items: center;
    justify-content: center;

    justify-self: center;
}}

.section-title {{
    font-size: 33px;
    font-weight: 900;
    line-height: 1.24;
    text-align: left;
}}

.section-title.tight {{
    font-size: 25px;
    line-height: 1.35;
    letter-spacing: -0.5px;
}}

.section-content {{
    padding: 22px 28px;
    font-size: 26px;
    line-height: 1.42;
    display: flex;
    flex-direction: column;
    justify-content: center;
}}

.thoughts {{
    margin: 0;
    padding-left: 28px;
    font-size: 29px;
    line-height: 1.52;
}}

.thoughts li {{
    margin-bottom: 8px;
}}

.table-row {{
    display: grid;
    grid-template-columns: 260px 1fr;
    border-bottom: 1px solid rgba(0,206,209,0.34);
    min-height: 42px;
}}

.table-row:last-child {{
    border-bottom: none;
}}

.stock-name {{
    color: #6ffcf4;
    font-size: 24px;
    font-weight: 900;
    display: flex;
    align-items: center;
    justify-content: center;
    border-right: 1px solid rgba(0,206,209,0.34);
    padding: 8px;
    text-align: center;
}}

.stock-desc {{
    font-size: 22px;
    display: flex;
    align-items: center;
    padding: 8px 16px;
    line-height: 1.28;
}}

.market-wrap {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 22px;
}}

.market-card {{
    position: relative;
    min-height: 154px;
    border: 1px solid rgba(0,206,209,0.45);
    border-radius: 12px;
    padding: 20px 26px;
    overflow: hidden;
}}

.market-name {{
    font-size: 34px;
    font-weight: 900;
}}

.market-value {{
    font-size: 52px;
    font-weight: 900;
    color: #ffffff;
    line-height: 1.1;
}}

.market-change {{
    font-size: 30px;
    font-weight: 900;
}}

.kospi-card .market-name,
.kospi-card .market-change {{
    color: #62e36d;
}}

.kosdaq-card .market-name,
.kosdaq-card .market-change {{
    color: #19c8ff;
}}

.market-graph-img {{
    position: absolute;
    right: 48px;
    top: 50%;
    transform: translateY(-50%);
    width: 105px;
    height: auto;
    object-fit: contain;
}}

.trade-row, .theme-row {{
    display: grid;
    grid-template-columns: 230px 1fr;
    border-bottom: 1px solid rgba(0,206,209,0.34);
}}

.trade-row:last-child, .theme-row:last-child {{
    border-bottom: none;
}}

.trade-type {{
    background: rgba(0,206,100,0.24);
    color: #99ffca;
    font-weight: 900;
    padding: 11px;
    text-align: center;
    border-right: 1px solid rgba(0,206,209,0.34);
    font-size: 22px;

    display: flex;
    align-items: center;
    justify-content: center;
}}

.trade-desc {{
    padding: 11px 18px;
    font-size: 22px;
    line-height: 1.35;

    display: flex;
    align-items: center;
}}

.theme-week {{
    background: rgba(255,128,0,0.16);
    color: #ffae4a;
    font-weight: 900;
    padding: 16px;
    text-align: center;
    border-right: 1px solid rgba(255,128,0,0.45);
    font-size: 24px;

    display: flex;
    align-items: center;
    justify-content: center;
}}

.theme-desc {{
    padding: 16px 20px;
    font-size: 23px;
    line-height: 1.34;
    font-weight: 800;
    display: block;
    white-space: normal;
}}

.theme-line {{
    display: block;
    margin-bottom: 6px;
    white-space: normal;
}}

.theme-label {{
    color: #ff9a28;
    font-weight: 900;
}}

.theme-text {{
    color: white;
    font-weight: 800;
}}
</style>
</head>

<body>
<div class="note">

    <div class="header">
        <div>
            <div class="date">{escape(note['date'])}</div>
            <div class="title">ClubQ NOTE</div>
        </div>
        <div class="link">litt.ly/clubq</div>
        <div class="slogan">지식과 자산이 쌓이는<br>투자커뮤니티 <span>ClubQ</span></div>
    </div>

    <div class="section">
        <div class="section-left">
            <div class="num">01</div>
            <div class="section-title">시장에<br>대한 생각</div>
        </div>
        <div class="section-content">
            <ul class="thoughts">{thoughts_html}</ul>
        </div>
    </div>

    <div class="section">
        <div class="section-left">
            <div class="num">02</div>
            <div class="section-title">상한가<br>등 특징주</div>
        </div>
        <div>{rows_stock(note['stocks'])}</div>
    </div>

    <div class="section">
        <div class="section-left">
            <div class="num">03</div>
            <div class="section-title">시장</div>
        </div>
        <div class="section-content market-wrap">
            <div class="market-card kospi-card">
                <div class="market-name">코스피</div>
                <div class="market-value">{escape(note['kospi']['value'])}</div>
                <div class="market-change">({escape(note['kospi']['change'])})</div>
                <img class="market-graph-img" src="{kospi_img}">
            </div>

            <div class="market-card kosdaq-card">
                <div class="market-name">코스닥</div>
                <div class="market-value">{escape(note['kosdaq']['value'])}</div>
                <div class="market-change">({escape(note['kosdaq']['change'])})</div>
                <img class="market-graph-img" src="{kosdaq_img}">
            </div>
        </div>
    </div>

    <div class="section">
        <div class="section-left">
            <div class="num">04</div>
            <div class="section-title">ClubQ<br>모임소식</div>
        </div>
        <div class="section-content">{clubq_html}</div>
    </div>

    <div class="section">
        <div class="section-left">
            <div class="num">05</div>
            <div class="section-title tight">호스트의<br>실제매매 및<br>관심기업(업종)</div>
        </div>
        <div>{rows_trades(note['trades'])}</div>
    </div>

    <div class="section">
        <div class="section-left">
            <div class="num">06</div>
            <div class="section-title tight">주도업종을<br>찾기 위한<br>업종흐름 파악</div>
        </div>
        <div>{rows_themes(note['themes'])}</div>
    </div>

</div>
</body>
</html>
"""


# =====================
# 파일 생성 / 이미지 캡처
# =====================

with open("test.html", "w", encoding="utf-8") as file:
    file.write(html_content)

with sync_playwright() as p:
    browser = p.chromium.launch()
    page = browser.new_page(viewport={"width": 1400, "height": 800}, device_scale_factor=1)
    page.goto(f"file://{os.getcwd()}/test.html")
    page.locator(".note").screenshot(path="clubq_note_final.png")
    browser.close()

print("ClubQ NOTE 최종 이미지 생성 완료!")
