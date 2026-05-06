import streamlit as st
import subprocess
import sys
import os
import re
from datetime import datetime
from PIL import Image
from openpyxl import load_workbook

# -----------------------------
# Playwright Chromium 설치
# -----------------------------
def ensure_playwright():
    subprocess.run(
        [sys.executable, "-m", "playwright", "install", "chromium"],
        check=False
    )

ensure_playwright()

# -----------------------------
# 파일명 생성 함수
# -----------------------------
def get_note_filename_from_excel(path="data.xlsx"):
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active

        raw = ws["A1"].value

        if not raw:
            return "ClubQ NOTE"

        first_line = str(raw).splitlines()[0].strip()

        # 예: 5.6.(화), 05.06.(화)
        match = re.search(r"(\d+)\.(\d+)", first_line)

        if not match:
            return "ClubQ NOTE"

        month = int(match.group(1))
        day = int(match.group(2))

        # 현재 연도 뒤 2자리
        year = datetime.now().strftime("%y")

        return f"{year}{month:02d}{day:02d} ClubQ NOTE"

    except Exception:
        return "ClubQ NOTE"

# -----------------------------
# Streamlit 페이지 설정
# -----------------------------
st.set_page_config(
    page_title="ClubQ NOTE Generator",
    layout="centered"
)

# -----------------------------
# 제목
# -----------------------------
st.title("📊 ClubQ NOTE 이미지 생성기")

st.write(
    "엑셀 파일을 업로드하면 NOTE 이미지를 자동 생성합니다."
)

# -----------------------------
# 엑셀 업로드
# -----------------------------
uploaded_file = st.file_uploader(
    "data.xlsx 업로드",
    type=["xlsx"]
)

# -----------------------------
# 업로드 성공 시
# -----------------------------
if uploaded_file:

    # 업로드 파일 저장
    with open("data.xlsx", "wb") as f:
        f.write(uploaded_file.getbuffer())

    st.success("✅ 엑셀 업로드 완료")

    # -----------------------------
    # 이미지 생성 버튼
    # -----------------------------
    if st.button("이미지 생성하기"):

        with st.spinner("이미지 생성 중입니다..."):

            result = subprocess.run(
                [sys.executable, "generate.py"],
                capture_output=True,
                text=True
            )

        # -----------------------------
        # 성공
        # -----------------------------
        if result.returncode == 0:

            st.success("✅ 이미지 생성 완료")

            image_path = "clubq_note_final.png"

            if os.path.exists(image_path):

                download_filename = f"{get_note_filename_from_excel('data.xlsx')}.png"

                image = Image.open(image_path)

                st.image(
                    image,
                    caption="생성된 ClubQ NOTE",
                    use_container_width=True
                )

                with open(image_path, "rb") as file:

                    st.download_button(
                        label="📥 이미지 다운로드",
                        data=file,
                        file_name=download_filename,
                        mime="image/png"
                    )

                st.caption(f"다운로드 파일명: {download_filename}")

            else:
                st.error("❌ 이미지 파일을 찾을 수 없습니다.")

        # -----------------------------
        # 실패
        # -----------------------------
        else:

            st.error("❌ 이미지 생성 실패")

            st.code(result.stderr)
